import csv
import io
from typing import List

from ..models.nota_fiscal import NotaFiscal

_CABECALHOS = [
    'Chave de Acesso', 'Número', 'Série', 'Data de Emissão',
    'CNPJ Emitente', 'Emitente', 'CNPJ Destinatário', 'Destinatário',
    'Valor Total (R$)', 'Tipo', 'Arquivo Original', 'Arquivo Organizado', 'Duplicata',
]


def _nota_para_linha(nota: NotaFiscal) -> list:
    return [
        nota.chave_acesso, nota.numero, nota.serie, nota.data_emissao,
        nota.emitente_cnpj, nota.emitente_nome,
        nota.destinatario_cnpj, nota.destinatario_nome,
        nota.valor_total, nota.tipo.upper(),
        nota.arquivo_original, nota.arquivo_organizado,
        'Sim' if nota.duplicata else 'Não',
    ]


def exportar_csv(notas: List[NotaFiscal]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(_CABECALHOS)
    for nota in notas:
        writer.writerow(_nota_para_linha(nota))
    return output.getvalue().encode('utf-8-sig')


def exportar_excel(notas: List[NotaFiscal]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return exportar_csv(notas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Notas Fiscais'

    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    dup_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for col, cabecalho in enumerate(_CABECALHOS, 1):
        cell = ws.cell(row=1, column=col, value=cabecalho)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, nota in enumerate(notas, 2):
        for col, valor in enumerate(_nota_para_linha(nota), 1):
            cell = ws.cell(row=row, column=col, value=valor)
            if nota.duplicata:
                cell.fill = dup_fill

    for col in ws.columns:
        largura = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(largura + 2, 60)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
